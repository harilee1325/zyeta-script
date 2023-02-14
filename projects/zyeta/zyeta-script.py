from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl 
from selenium.webdriver.support.ui import Select


import time

PATH = "/Users/harikrishnan/projects/zyeta/chromedriver"
driver = webdriver.Chrome(PATH)


## Adding main question

def loginUser():
    driver.get("https://kpi.zyeta.com/kpi/login")
    print(driver.title)
    ## Logging in user
    email = driver.find_element(By.ID, "email" )
    email.send_keys("shafeeq@zyeta.com")

    password = driver.find_element(By.ID, "password")
    password.send_keys("123456789")

    loginButton = driver.find_element(By.XPATH, '//*[@type="submit"]')
    loginButton.click()


def goToCreateKpi():
    ## redirecting and creating a form
    createKpi = driver.find_element(By.XPATH, '//*[@id="side-bar"]/nav/ul/li[1]/a')
    createKpi.click()

    kpiTitle = driver.find_element(By.XPATH, '//*[@name="title"]')
    kpiTitle.send_keys("Growth")

    time.sleep(2)
    createKpiBt = driver.find_element(By.XPATH, '//*[@type="submit"]')
    createKpiBt.click()

   

def addSubQuestion(sub_data, sheet_count):
    
    for data in sub_data:
        saveQ = driver.find_elements(By.XPATH, '//*[text()="Add Sub-question"]')
        saveQ[sheet_count].click()
        time.sleep(3)
        kpiDesc = driver.find_element(By.XPATH, '//*[@name="qnDesc"]')
        kpiDesc.send_keys(data.getTitle())
        time.sleep(2)

        q_val = ""
        if (data.getQtype() == 1):
            q_val = "multiple"
        elif (data.getQtype() == 2):
            q_val = "yes"
        else:
            q_val = "numerical"

        select = Select(driver.find_element(By.ID, 'typeSelector'))
        select.select_by_value(q_val)

        if (data.getQtype() == 1):
            numberOfChoice = driver.find_element(By.ID, 'multipleNumber')
            numberOfChoice.send_keys(data.getChoices())
            time.sleep(2)
            hasDocTxt = driver.find_element(By.CLASS_NAME, 'checkbox')
            hasDocTxt.click()
            time.sleep(2)

            if (data.getHasDoc() == "Y"):
                hasDoc = driver.find_element(By.NAME, 'document')
                hasDoc.click()
            
            for i in range(1, len(data.getChoices())+1):
                ans = driver.find_element(By.NAME, 'ans_'+str(i))
                ans.send_keys(data.getChoices()[i-1])
                time.sleep(2)
                point = driver.find_element(By.NAME, 'ans_point_'+str(i))
                point.send_keys(data.getChoiceMarks()[i-1])

        elif (data.getQtype() == 2):
            time.sleep(3)
            if (data.getHasDoc() == "Y"):
                hasDoc = driver.find_element(By.NAME, 'document')
                hasDoc.click()

            print(data.getChoiceMarks())
            yesPoint = driver.find_element(By.NAME, 'yesPoints')
            yesPoint.send_keys(data.getChoiceMarks()[0])

            noPoint = driver.find_element(By.NAME, 'noPoints')
            noPoint.send_keys(data.getChoiceMarks()[1])


        saveQ = driver.find_element(By.NAME, 'submit')
        saveQ.click()
        time.sleep(3)



def addKPI(title, desc, q_type, choice, marks, choice_title, has_document, sub_data, sheet_count):
     ## clicking on add question button
    time.sleep(2)
    addQ = driver.execute_script('return document.querySelector("body div > a.button.is-secondary")')
    addQ.click()

    kpiTitle = driver.find_element(By.XPATH, '//*[@name="qnTitle"]')
    kpiTitle.send_keys(title)
    time.sleep(3)

    kpiDesc = driver.find_element(By.XPATH, '//*[@name="qnDesc"]')
    kpiDesc.send_keys(desc)
    time.sleep(2)

    
    q_val = ""
    if (q_type == 1):
        q_val = "multiple"
    elif (q_type == 2):
        q_val = "yes"
    else:
        q_val = "numerical"

    select = Select(driver.find_element(By.ID, 'typeSelector'))
    select.select_by_value(q_val)

    if (q_type == 1):
        numberOfChoice = driver.find_element(By.ID, 'multipleNumber')
        numberOfChoice.send_keys(choice)
        time.sleep(2)
        hasDocTxt = driver.find_element(By.CLASS_NAME, 'checkbox')
        hasDocTxt.click()
        time.sleep(2)

        if (has_document == "Y"):
            hasDoc = driver.find_element(By.NAME, 'document')
            hasDoc.click()
        
        for i in range(1, len(choice_title)+1):
            ans = driver.find_element(By.NAME, 'ans_'+str(i))
            ans.send_keys(choice_title[i-1])
            time.sleep(2)
            point = driver.find_element(By.NAME, 'ans_point_'+str(i))
            point.send_keys(marks[i-1])

    elif (q_type == 2):
        time.sleep(3)
        if (has_document == "Y"):
            hasDoc = driver.find_element(By.NAME, 'document')
            hasDoc.click()

        print(marks)
        yesPoint = driver.find_element(By.NAME, 'yesPoints')
        yesPoint.send_keys(marks[0])

        noPoint = driver.find_element(By.NAME, 'noPoints')
        noPoint.send_keys(marks[1])

    else:
        time.sleep(3)
        if (has_document == "Y"):
            hasDoc = driver.find_element(By.NAME, 'document')
            hasDoc.click()

        print(marks)
        yesPoint = driver.find_element(By.NAME, 'numericalTarget')
        yesPoint.send_keys(marks[0])

        noPoint = driver.find_element(By.NAME, 'maximumNumericalPoint')
        noPoint.send_keys(marks[1])


    saveQ = driver.find_element(By.NAME, 'submit')
    saveQ.click()
    time.sleep(3)
    addSubQuestion(sub_data, sheet_count)

   # if (sub_data == 0 or len(sub_data) != 0):
       
       


time.sleep(2)

loginUser()
goToCreateKpi()




class Data:
    def __init__(self, title, description, q_type, marks, has_document, sub_data, choices, choice_marks, sheet_count):
        self.title = title
        self.description = description
        self.q_type = q_type
        self.marks = marks
        self.has_document = has_document
        self.sub_data = sub_data
        self.choices = choices
        self.choice_marks = choice_marks
        self.sheet_count = sheet_count

    def getCount(self):
        return self.sheet_count

    def getTitle(self):
        return self.title

    def getDesc(self):
        return self.description

    def getQtype(self):
        return self.q_type

    def getMarks(self):
        return self.marks

    def getHasDoc(self):
        return self.has_document
    
    def getSubData(self):
        return self.sub_data

    def getChoices(self):
        return self.choices

    def getChoiceMarks(self):
        return str(self.choice_marks).split(",")


    def printAll(self):
        print(self.title)
        print(self.description)
        print(self.q_type)
        print(self.marks)
        print(self.has_document)
        print(self.choices)
        print(self.choice_marks)
        print("printing sub data")
        for subData in self.sub_data:
            subData.printAll()

        

class SubData:
    def __init__(self, title, q_type, marks, has_document, choices, choice_marks):
        self.title = title
        self.q_type = q_type
        self.marks = marks
        self.has_document = has_document
        self.choices = choices
        self.choice_marks = choice_marks

    def getTitle(self):
        return self.title

    def getQtype(self):
        return self.q_type

    def getMarks(self):
        return self.marks

    def getHasDoc(self):
        return self.has_document

    def getChoices(self):
        return self.choices

    def getChoiceMarks(self):
        return self.choice_marks.split(",")

    def printAll(self):
        print(self.title)
        print(self.q_type)
        print(self.marks)
        print(self.has_document)
        print(self.choices)
        print(self.choice_marks)


data_path = "/Users/harikrishnan/projects/zyeta/data.xlsx"

wb_obj = openpyxl.load_workbook(data_path) 

#sheet_obj = wb_obj.active 
sheet_count = 0

list = []
sub_list = []

for sheet in wb_obj.worksheets:

    row = sheet.max_row
    column = sheet.max_column


    
    title_val = ""
    desc_val = ""
    q_type_val = ""
    marks_val = ""
    choice_title_val = []
    choice_mark_val = ""
    has_document_val = ""

    sub_choice_title = []
    sub_choice_value = ""

    for i in range(2, row + 1): 
        cell_obj = sheet.cell(row = i, column = 1) 
        if (cell_obj.value == None):
            break
        title = sheet.cell(row = i, column=1)
        desc = sheet.cell(row = i, column=2)
        q_type = sheet.cell(row = i, column=3)
        mark = sheet.cell(row = i, column=4)
        choice_title = sheet.cell(row = i, column=5)
        choice_mark = sheet.cell(row = i, column=6)
        has_document = sheet.cell(row = i, column=7)

        # print(title.value)
        # print(desc.value)
        # print(q_type.value)
        if (i == 2):     
            title_val = title.value
            desc_val = desc.value
            q_type_val = q_type.value
            marks_val = mark.value
            has_document_val = has_document.value
            if (choice_title.value != None):
                choice_title_val = choice_title.value.split(",")
            if (choice_mark.value != None):
                choice_mark_val = choice_mark.value
            else:
                choice_mark_val = 0

        else:
            if (choice_title.value != None):
                sub_choice_title = choice_title.value.split(",")
            if (choice_mark.value != None):
                sub_choice_value = choice_mark.value
            else:
                sub_choice_value = 0
            sub_list.append(SubData(title=title.value, q_type=q_type.value, marks= mark.value, has_document=has_document.value, choices= sub_choice_title, choice_marks=sub_choice_value))


    list.append(Data(title=title_val
    , description=desc_val
    , q_type=q_type_val
    , marks= marks_val
    , has_document=has_document_val
    , sub_data=sub_list
    , choices= choice_title_val
    , choice_marks=choice_mark_val
    , sheet_count = sheet_count))
    sub_list = []
    sheet_count = sheet_count + 1

    


print(len(list))
for data in list:
    print(data.getTitle())
    addKPI( data.getTitle(), data.getDesc(), data.getQtype(), len(data.getChoices()), data.getChoiceMarks(),data.getChoices(), data.getHasDoc(), data.getSubData(), data.getCount())


time.sleep(1000000)